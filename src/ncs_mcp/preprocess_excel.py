from __future__ import annotations

if __package__ in {None, ""}:  # pragma: no cover - direct script support
    import sys
    from pathlib import Path

    sys.path.append(str(Path(__file__).resolve().parents[1]))

import argparse
import json
import sqlite3
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ncs_mcp.config import load_settings
from ncs_mcp.db import (
    connect,
    create_indexes,
    initialize_database,
    now_utc,
    parse_element_no,
    split_unit_code,
)


HEADER_ALIASES = {
    "major_code": ("대분류코드",),
    "major_name": ("대분류코드명", "대분류명"),
    "middle_code": ("중분류코드",),
    "middle_name": ("중분류코드명", "중분류명"),
    "small_code": ("소분류코드",),
    "small_name": ("소분류코드명", "소분류명"),
    "sub_code": ("세분류코드",),
    "sub_name": ("세분류코드명", "세분류명"),
    "unit_code": ("능력단위분류번호", "능력단위코드"),
    "unit_name": ("능력단위명칭", "능력단위명"),
    "unit_level": ("수준", "능력단위수준"),
    "element_code": ("능력단위요소번호",),
    "element_name": ("능력단위요소명", "능력단위요소"),
    "element_level": ("능력단위요소수준",),
    "criteria_no": ("수행준거번호",),
    "criteria_text": ("수행준거",),
    "ksa_type_code": ("지식기술태도코드",),
    "ksa_type_name": ("지식기술태도코드명",),
    "ksa_no": ("지식기술태도번호",),
    "ksa_text": ("지식기술태도의의", "지식기술태도"),
}


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def build_header_map(header: Iterable[Any]) -> dict[str, int]:
    normalized = {text(name): idx for idx, name in enumerate(header)}
    result: dict[str, int] = {}
    missing: list[str] = []
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                result[field] = normalized[alias]
                break
        else:
            missing.append(field)
    if missing:
        raise ValueError(f"Missing required NCS columns: {', '.join(missing)}")
    return result


def get(row: tuple[Any, ...], header_map: dict[str, int], field: str) -> str:
    idx = header_map[field]
    return text(row[idx]) if idx < len(row) else ""


class Normalizer:
    def __init__(self, conn: sqlite3.Connection, source_file: str):
        self.conn = conn
        self.source_file = source_file
        self.classification_cache: dict[tuple[str, str, str, str], int] = {}
        self.element_cache: dict[tuple[str, str, str], int] = {}
        self.criteria_cache: dict[tuple[int, str, str], int] = {}
        self.ksa_cache: dict[tuple[int, str, str, str], int] = {}

    def get_classification_id(self, values: dict[str, str]) -> int:
        key = (
            values["major_code"],
            values["middle_code"],
            values["small_code"],
            values["sub_code"],
        )
        cached = self.classification_cache.get(key)
        if cached:
            return cached
        self.conn.execute(
            """
            INSERT OR IGNORE INTO classifications(
                major_code, major_name, middle_code, middle_name,
                small_code, small_name, sub_code, sub_name
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                values["major_code"],
                values["major_name"],
                values["middle_code"],
                values["middle_name"],
                values["small_code"],
                values["small_name"],
                values["sub_code"],
                values["sub_name"],
            ),
        )
        row = self.conn.execute(
            """
            SELECT classification_id
            FROM classifications
            WHERE major_code = ? AND middle_code = ? AND small_code = ? AND sub_code = ?
            """,
            key,
        ).fetchone()
        classification_id = int(row["classification_id"])
        self.classification_cache[key] = classification_id
        return classification_id

    def upsert_unit(self, values: dict[str, str], classification_id: int, timestamp: str) -> None:
        base_code, version = split_unit_code(values["unit_code"])
        self.conn.execute(
            """
            INSERT INTO competency_units(
                unit_code, base_unit_code, unit_version, unit_name_raw,
                unit_level_raw, classification_id, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(unit_code) DO UPDATE SET
                unit_name_raw = excluded.unit_name_raw,
                unit_level_raw = excluded.unit_level_raw,
                classification_id = excluded.classification_id,
                updated_at = excluded.updated_at
            """,
            (
                values["unit_code"],
                base_code,
                version,
                values["unit_name"],
                values["unit_level"],
                classification_id,
                timestamp,
                timestamp,
            ),
        )

    def get_element_id(self, values: dict[str, str]) -> int:
        element_no = parse_element_no(values["element_code"], values["unit_code"])
        key = (values["unit_code"], values["element_code"], values["element_name"])
        cached = self.element_cache.get(key)
        if cached:
            return cached
        self.conn.execute(
            """
            INSERT OR IGNORE INTO competency_elements(
                unit_code, element_no, element_code_raw, element_name_raw, element_level_raw
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (
                values["unit_code"],
                element_no,
                values["element_code"],
                values["element_name"],
                values["element_level"],
            ),
        )
        row = self.conn.execute(
            """
            SELECT element_id
            FROM competency_elements
            WHERE unit_code = ? AND element_code_raw = ? AND element_name_raw = ?
            """,
            key,
        ).fetchone()
        element_id = int(row["element_id"])
        self.element_cache[key] = element_id
        return element_id

    def get_criteria_id(self, element_id: int, values: dict[str, str]) -> int:
        key = (element_id, values["criteria_no"], values["criteria_text"])
        cached = self.criteria_cache.get(key)
        if cached:
            return cached
        self.conn.execute(
            """
            INSERT OR IGNORE INTO performance_criteria(
                element_id, criteria_no, criteria_text_raw
            ) VALUES (?, ?, ?)
            """,
            (element_id, values["criteria_no"], values["criteria_text"]),
        )
        row = self.conn.execute(
            """
            SELECT criteria_id
            FROM performance_criteria
            WHERE element_id = ? AND criteria_no = ? AND criteria_text_raw = ?
            """,
            key,
        ).fetchone()
        criteria_id = int(row["criteria_id"])
        self.criteria_cache[key] = criteria_id
        return criteria_id

    def get_ksa_id(self, element_id: int, values: dict[str, str]) -> int:
        key = (
            element_id,
            values["ksa_type_code"],
            values["ksa_no"],
            values["ksa_text"],
        )
        cached = self.ksa_cache.get(key)
        if cached:
            return cached
        self.conn.execute(
            """
            INSERT OR IGNORE INTO ksa_items(
                element_id, ksa_type_code, ksa_type_name, ksa_no, ksa_text_raw
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (
                element_id,
                values["ksa_type_code"],
                values["ksa_type_name"],
                values["ksa_no"],
                values["ksa_text"],
            ),
        )
        row = self.conn.execute(
            """
            SELECT ksa_id
            FROM ksa_items
            WHERE element_id = ? AND ksa_type_code = ? AND ksa_no = ? AND ksa_text_raw = ?
            """,
            key,
        ).fetchone()
        ksa_id = int(row["ksa_id"])
        self.ksa_cache[key] = ksa_id
        return ksa_id

    def insert_raw_row(self, sheet_name: str, sheet_row_number: int, values: dict[str, str], timestamp: str) -> int:
        cursor = self.conn.execute(
            """
            INSERT INTO raw_excel_rows(
                source_file, sheet_name, sheet_row_number,
                major_code, major_name, middle_code, middle_name,
                small_code, small_name, sub_code, sub_name,
                unit_code, unit_name, unit_level,
                element_code, element_name, element_level,
                criteria_no, criteria_text,
                ksa_type_code, ksa_type_name, ksa_no, ksa_text,
                loaded_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                self.source_file,
                sheet_name,
                sheet_row_number,
                values["major_code"],
                values["major_name"],
                values["middle_code"],
                values["middle_name"],
                values["small_code"],
                values["small_name"],
                values["sub_code"],
                values["sub_name"],
                values["unit_code"],
                values["unit_name"],
                values["unit_level"],
                values["element_code"],
                values["element_name"],
                values["element_level"],
                values["criteria_no"],
                values["criteria_text"],
                values["ksa_type_code"],
                values["ksa_type_name"],
                values["ksa_no"],
                values["ksa_text"],
                timestamp,
            ),
        )
        return int(cursor.lastrowid)

    def insert_link(self, raw_row_id: int, element_id: int, criteria_id: int, ksa_id: int) -> None:
        self.conn.execute(
            """
            INSERT OR IGNORE INTO element_criteria_ksa_links(
                raw_row_id, element_id, criteria_id, ksa_id
            ) VALUES (?, ?, ?, ?)
            """,
            (raw_row_id, element_id, criteria_id, ksa_id),
        )

    def ingest(self, sheet_name: str, sheet_row_number: int, values: dict[str, str], timestamp: str) -> None:
        classification_id = self.get_classification_id(values)
        self.upsert_unit(values, classification_id, timestamp)
        element_id = self.get_element_id(values)
        criteria_id = self.get_criteria_id(element_id, values)
        ksa_id = self.get_ksa_id(element_id, values)
        raw_row_id = self.insert_raw_row(sheet_name, sheet_row_number, values, timestamp)
        self.insert_link(raw_row_id, element_id, criteria_id, ksa_id)


def reset_database(db_path: Path) -> None:
    for suffix in ("", "-wal", "-shm"):
        path = Path(str(db_path) + suffix)
        if path.exists():
            path.unlink()


def summarize(conn: sqlite3.Connection, source_file: str, sheets: list[dict[str, Any]]) -> dict[str, Any]:
    tables = {
        "raw_excel_rows": "원본 엑셀 행",
        "classifications": "세분류",
        "competency_units": "능력단위",
        "competency_elements": "능력단위요소",
        "performance_criteria": "수행준거",
        "ksa_items": "KSA",
        "element_criteria_ksa_links": "원행 조합 링크",
    }
    counts = {
        table: int(conn.execute(f"SELECT COUNT(*) AS count FROM {table}").fetchone()["count"])
        for table in tables
    }
    return {
        "source_file": source_file,
        "generated_at": now_utc(),
        "counts": counts,
        "sheets": sheets,
    }


def write_reports(summary: dict[str, Any], reports_dir: Path) -> None:
    reports_dir.mkdir(parents=True, exist_ok=True)
    json_path = reports_dir / "preprocess_summary.json"
    md_path = reports_dir / "preprocess_summary.md"
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# NCS 전처리 요약",
        "",
        f"- 생성시각: {summary['generated_at']}",
        f"- 원본파일: {summary['source_file']}",
        "",
        "## 테이블별 건수",
        "",
        "| 테이블 | 건수 |",
        "|---|---:|",
    ]
    for table, count in summary["counts"].items():
        lines.append(f"| `{table}` | {count:,} |")
    lines.extend(["", "## 시트별 처리 행", "", "| 시트 | 처리 행 |", "|---|---:|"])
    for sheet in summary["sheets"]:
        lines.append(f"| {sheet['sheet_name']} | {sheet['rows']:,} |")
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def preprocess_excel(
    excel_path: Path,
    db_path: Path,
    reports_dir: Path,
    reset: bool = False,
    sheets: set[str] | None = None,
    max_rows: int | None = None,
    commit_interval: int = 5000,
) -> dict[str, Any]:
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    if reset:
        reset_database(db_path)

    conn = connect(db_path)
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = NORMAL")
    initialize_database(conn)

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    normalizer = Normalizer(conn, excel_path.name)
    processed_sheets: list[dict[str, Any]] = []
    total_rows = 0
    timestamp = now_utc()

    try:
        for worksheet in workbook.worksheets:
            if sheets and worksheet.title not in sheets:
                continue
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            header_map = build_header_map(header)
            sheet_rows = 0

            with conn:
                for row_number, row in enumerate(rows, start=2):
                    values = {field: get(row, header_map, field) for field in HEADER_ALIASES}
                    normalizer.ingest(worksheet.title, row_number, values, timestamp)
                    sheet_rows += 1
                    total_rows += 1
                    if total_rows % commit_interval == 0:
                        conn.commit()
                    if max_rows is not None and total_rows >= max_rows:
                        break
            processed_sheets.append({"sheet_name": worksheet.title, "rows": sheet_rows})
            if max_rows is not None and total_rows >= max_rows:
                break
    finally:
        workbook.close()

    create_indexes(conn)
    summary = summarize(conn, excel_path.name, processed_sheets)
    write_reports(summary, reports_dir)
    conn.close()
    return summary


def parse_args() -> argparse.Namespace:
    settings = load_settings()
    parser = argparse.ArgumentParser(description="Preprocess NCS Excel DB into SQLite.")
    parser.add_argument("--excel-path", type=Path, default=settings.excel_path)
    parser.add_argument("--db-path", type=Path, default=settings.db_path)
    parser.add_argument("--reports-dir", type=Path, default=settings.reports_dir)
    parser.add_argument("--reset", action="store_true")
    parser.add_argument("--sheets", help="Comma-separated sheet names, e.g. 02,03")
    parser.add_argument("--max-rows", type=int)
    parser.add_argument("--commit-interval", type=int, default=5000)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.excel_path is None:
        raise SystemExit("NCS_EXCEL_PATH is required. Set it in .env or pass --excel-path.")
    selected_sheets = {part.strip() for part in args.sheets.split(",")} if args.sheets else None
    summary = preprocess_excel(
        excel_path=args.excel_path,
        db_path=args.db_path,
        reports_dir=args.reports_dir,
        reset=args.reset,
        sheets=selected_sheets,
        max_rows=args.max_rows,
        commit_interval=args.commit_interval,
    )
    print(json.dumps(summary["counts"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
